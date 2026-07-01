"""
APP DE CONVERSIÓN XML A PDF - TICKET 72mm
"""

import xml.etree.ElementTree as ET
from fpdf import FPDF
import os
import tempfile
import io
from typing import Dict, Any, Optional
import base64
from datetime import datetime
import qrcode
import logging
import uuid
from functools import wraps
import openpyxl
import requests
from flask import Flask, request, send_file, render_template_string, jsonify, session, redirect, url_for
from flask_session import Session

# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'manchester_pro_secret_key_2024')

# Configurar sesión con Redis para producción
redis_url = os.environ.get('REDIS_URL', None)
if redis_url:
    try:
        import redis
        app.config['SESSION_TYPE'] = 'redis'
        app.config['SESSION_PERMANENT'] = False
        app.config['SESSION_USE_SIGNER'] = True
        app.config['SESSION_KEY_PREFIX'] = 'manchester:'
        app.config['SESSION_REDIS'] = redis.from_url(redis_url)
        logger.info(f"Redis configurado correctamente: {redis_url[:20]}...")
    except Exception as e:
        logger.error(f"Error configurando Redis: {e}")
        app.config['SESSION_TYPE'] = 'filesystem'
else:
    app.config['SESSION_TYPE'] = 'filesystem'
    app.config['SESSION_FILE_DIR'] = 'temp_sessions'
    logger.info("Redis no configurado, usando sesión por defecto (filesystem)")

# Crear directorio de sesiones si es filesystem
if app.config.get('SESSION_TYPE') == 'filesystem':
    os.makedirs('temp_sessions', exist_ok=True)

# Inicializar Flask-Session
Session(app)

# Configurar sesión para producción
app.config.update(
    SESSION_COOKIE_SECURE=False,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=3600  # 1 hora
) 

# Configuración de la app
CONFIG = {
    'MAX_FILE_SIZE': 10 * 1024 * 1024,  # 10MB max
    'ALLOWED_EXTENSIONS': ['.xml', '.csv', '.xlsx'],
    'DEFAULT_FORMAT': 'ticket',
    'PAGE_WIDTH': 72,
}

YAPES_SHEETS_URL = "https://docs.google.com/spreadsheets/d/1-egUiQ0K7vYh1rqT0EC0acx5zbFzy5V6IoEGwrNPiAs/export?format=csv"



class YapesPDF:
    """Clase para generar PDF de YAPES"""
    
    def __init__(self, csv_path: str = '', output_path: str = '', fecha_inicio: str = '', fecha_fin: str = '', sheets_url: str = ''):
        self.file_path = csv_path
        self.output_path = output_path
        self.sheets_url = sheets_url
        self.data = []
        self.rango_fechas_yape = ""
        self.fecha_inicio = fecha_inicio
        self.fecha_fin = fecha_fin
        logger.info(f"YapesPDF INIT: fecha_inicio={fecha_inicio}, fecha_fin={fecha_fin}, sheets_url={'SI' if sheets_url else 'NO'}")
        
    def _limpiar_texto(self, texto: str) -> str:
        """Limpiar texto: quitar espacios extra, comillas, etc."""
        if not texto:
            return ""
        texto = texto.strip()
        texto = texto.strip('"').strip("'").strip()
        import re
        texto = re.sub(r'\s+', ' ', texto)
        return texto
    
    def _normalizar_nombre(self, nombre: str) -> str:
        """Normalizar nombre para agrupar: mayúsculas, sin tildes, sin acentos"""
        import unicodedata
        nombre = nombre.upper().strip()
        # Eliminar tildes/acentos
        nombre_normalizado = unicodedata.normalize('NFD', nombre)
        nombre_normalizado = ''.join(c for c in nombre_normalizado if unicodedata.category(c) != 'Mn')
        return nombre_normalizado
    
    def _parsear_monto(self, raw: str) -> float:
        """Parsear monto: soporta '128.5 blabla' → 128.5 y '128 con 6' → 128.60"""
        raw = raw.strip()
        import re
        # Detectar patrón "X con Y" (ej: "128 con 6" → 128.60)
        m = re.search(r'(\d+)\s*con\s*(\d+)', raw, re.IGNORECASE)
        if m:
            entero = m.group(1)
            decimal = m.group(2).ljust(2, '0')[:2]
            return float(f"{entero}.{decimal}")
        # Caso normal: limpiar y tomar número del inicio
        s = raw.replace(',', '.')
        s = re.sub(r'[^\d.]', '', s)
        try:
            return float(s)
        except ValueError:
            return 0.0
    
    def _parsear_fecha(self, fecha_raw: str) -> tuple:
        """Parsear fecha, separar fecha y hora si existe"""
        fecha_raw = self._limpiar_texto(fecha_raw)
        
        if not fecha_raw:
            fecha = datetime.now().strftime('%d/%m/%Y')
            hora = ""
            return fecha, hora
        
        # Intentar diferentes formatos (año 2 dígitos y 4 dígitos)
        formatos = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%d',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y',
            '%d/%m/%y',  # Año 2 dígitos
            '%d-%m-%y',
        ]
        
        fecha = ""
        hora = ""
        
        for fmt in formatos:
            try:
                dt = datetime.strptime(fecha_raw, fmt)
                # Convertir a DD/MM/YYYY para consistencia
                fecha = dt.strftime('%d/%m/%Y')
                if dt.hour > 0 or dt.minute > 0 or dt.second > 0:
                    hora = dt.strftime('%H:%M')
                break
            except:
                continue
        
        if not fecha:
            logger.warning(f"No se pudo parsear fecha: {fecha_raw}")
            fecha = datetime.now().strftime('%d/%m/%Y')
        
        return fecha, hora
        
    def parse_csv(self) -> bool:
        """Parsear archivo CSV"""
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            logger.info(f"CSV leído: total líneas={len(lines)}, primera={lines[1][:50] if len(lines) > 1 else 'N/A'}, tamaño_archivo={os.path.getsize(self.file_path) if os.path.exists(self.file_path) else 'N/A'} bytes")
            
            if len(lines) < 2:
                return False
            
            lineas_procesadas = 0
            datos_anadidos = 0
            
            for line in lines[1:]:
                lineas_procesadas += 1
                if not line.strip():
                    continue
                
                # Separar por coma, manejar comillas dentro de valores
                import re
                parts = []
                in_quote = False
                current = ""
                for char in line.strip():
                    if char == '"':
                        in_quote = not in_quote
                    elif char == ',' and not in_quote:
                        parts.append(current)
                        current = ""
                    else:
                        current += char
                parts.append(current)
                
                # Limpiar comillas y espacios
                parts = [self._limpiar_texto(p) for p in parts]
                logger.info(f"Línea {lineas_procesadas}: {len(parts)} partes, parts={parts[:3]}")
                
                if len(parts) >= 2 and parts[0] and parts[1]:
                    monto = self._parsear_monto(parts[1])
                    
                    # Parsear fecha y hora
                    fecha_raw = parts[2].strip() if len(parts) > 2 else ""
                    fecha, hora = self._parsear_fecha(fecha_raw)
                    
                    # Filtrar por rango de fechas si se especificó
                    if self.fecha_inicio or self.fecha_fin:
                        try:
                            item_date = datetime.strptime(fecha, '%d/%m/%Y')
                            if self.fecha_inicio:
                                inicio = datetime.strptime(self.fecha_inicio, '%Y-%m-%d')
                                if item_date < inicio:
                                    continue
                            if self.fecha_fin:
                                fin = datetime.strptime(self.fecha_fin, '%Y-%m-%d')
                                if item_date > fin:
                                    continue
                        except Exception as e:
                            logger.warning(f"Error filtrando fecha '{fecha}': {e}")
                            pass
                    
                    tokens = parts[0].split()
                    primer_nombre = tokens[0] if tokens else parts[0]
                    nombre_original = primer_nombre.upper()
                    nombre_normalizado = self._normalizar_nombre(nombre_original)
                    
                    self.data.append({
                        'nombre': nombre_original,
                        'nombre_key': nombre_normalizado,
                        'monto': monto,
                        'fecha': fecha,
                        'hora': hora
                    })
                    datos_anadidos += 1
                    logger.info(f"Datos añadidos: {datos_anadidos}, total: {len(self.data)}")
            
            logger.info(f"CSV parseado: líneas={lineas_procesadas}, añadidos={datos_anadidos}, data={len(self.data)}")
            return len(self.data) > 0
        except Exception as e:
            logger.error(f"Error parseando CSV: {e}")
            return False

    def parse_xlsx(self) -> bool:
        """Parsear archivo XLSX"""
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            ws = wb.active
            if ws is None:
                logger.error("No se encontró hoja activa en el XLSX")
                return False

            filas = list(ws.iter_rows(values_only=True))
            if len(filas) < 2:
                logger.warning("XLSX sin suficientes filas (mínimo 2)")
                return False

            logger.info(f"XLSX leído: total filas={len(filas)}")

            lineas_procesadas = 0
            datos_anadidos = 0

            for row in filas[1:]:
                lineas_procesadas += 1
                if not row or all(v is None for v in row):
                    continue

                parts = [str(v).strip() if v is not None else '' for v in row]
                logger.info(f"Fila {lineas_procesadas}: {len(parts)} partes, parts={parts[:3]}")

                if len(parts) >= 2 and parts[0] and parts[1]:
                    monto = self._parsear_monto(parts[1])

                    if monto == 0.0:
                        logger.warning(f"Monto inválido: {parts[1]}, usando 0")
                        monto = 0.0

                    fecha_raw = parts[2].strip() if len(parts) > 2 else ""
                    fecha, hora = self._parsear_fecha(fecha_raw)

                    if self.fecha_inicio or self.fecha_fin:
                        try:
                            item_date = datetime.strptime(fecha, '%d/%m/%Y')
                            if self.fecha_inicio:
                                inicio = datetime.strptime(self.fecha_inicio, '%Y-%m-%d')
                                if item_date < inicio:
                                    continue
                            if self.fecha_fin:
                                fin = datetime.strptime(self.fecha_fin, '%Y-%m-%d')
                                if item_date > fin:
                                    continue
                        except Exception as e:
                            logger.warning(f"Error filtrando fecha '{fecha}': {e}")
                            pass

                    tokens = parts[0].split()
                    primer_nombre = tokens[0] if tokens else parts[0]
                    nombre_original = primer_nombre.upper()
                    nombre_normalizado = self._normalizar_nombre(nombre_original)

                    self.data.append({
                        'nombre': nombre_original,
                        'nombre_key': nombre_normalizado,
                        'monto': monto,
                        'fecha': fecha,
                        'hora': hora
                    })
                    datos_anadidos += 1

            logger.info(f"XLSX parseado: filas={lineas_procesadas}, añadidos={datos_anadidos}, data={len(self.data)}")
            return len(self.data) > 0
        except Exception as e:
            logger.error(f"Error parseando XLSX: {e}")
            return False

    def parse_gsheets(self) -> bool:
        """Parsear datos desde Google Sheets (export CSV)"""
        try:
            logger.info(f"Descargando datos desde Google Sheets: {self.sheets_url[:60]}...")
            resp = requests.get(self.sheets_url, timeout=30)
            resp.raise_for_status()
            content = resp.text

            lines = content.splitlines()
            if len(lines) < 2:
                logger.warning("Google Sheets sin suficientes filas")
                return False

            logger.info(f"Google Sheets: {len(lines)} filas descargadas")

            lineas_procesadas = 0
            datos_anadidos = 0

            for line in lines[1:]:
                lineas_procesadas += 1
                if not line.strip():
                    continue

                parts = []
                in_quote = False
                current = ""
                for char in line.strip():
                    if char == '"':
                        in_quote = not in_quote
                    elif char == ',' and not in_quote:
                        parts.append(current)
                        current = ""
                    else:
                        current += char
                parts.append(current)

                parts = [self._limpiar_texto(p) for p in parts]

                if len(parts) >= 2 and parts[0] and parts[1]:
                    monto = self._parsear_monto(parts[1])

                    if monto == 0.0:
                        logger.warning(f"Monto inválido: {parts[1]}, usando 0")
                        monto = 0.0

                    fecha_raw = parts[2].strip() if len(parts) > 2 else ""
                    fecha, hora = self._parsear_fecha(fecha_raw)

                    if self.fecha_inicio or self.fecha_fin:
                        try:
                            item_date = datetime.strptime(fecha, '%d/%m/%Y')
                            if self.fecha_inicio:
                                inicio = datetime.strptime(self.fecha_inicio, '%Y-%m-%d')
                                if item_date < inicio:
                                    continue
                            if self.fecha_fin:
                                fin = datetime.strptime(self.fecha_fin, '%Y-%m-%d')
                                if item_date > fin:
                                    continue
                        except Exception as e:
                            logger.warning(f"Error filtrando fecha '{fecha}': {e}")
                            pass

                    tokens = parts[0].split()
                    primer_nombre = tokens[0] if tokens else parts[0]
                    nombre_original = primer_nombre.upper()
                    nombre_normalizado = self._normalizar_nombre(nombre_original)

                    self.data.append({
                        'nombre': nombre_original,
                        'nombre_key': nombre_normalizado,
                        'monto': monto,
                        'fecha': fecha,
                        'hora': hora
                    })
                    datos_anadidos += 1

            logger.info(f"Google Sheets parseado: filas={lineas_procesadas}, añadidos={datos_anadidos}, data={len(self.data)}")
            return len(self.data) > 0
        except requests.RequestException as e:
            logger.error(f"Error de red al descargar Google Sheets: {e}")
            return False
        except Exception as e:
            logger.error(f"Error parseando Google Sheets: {e}")
            return False

    def parse(self) -> bool:
        """Auto-detectar fuente de datos y parsear"""
        if self.sheets_url:
            return self.parse_gsheets()
        if self.file_path.lower().endswith('.xlsx'):
            return self.parse_xlsx()
        return self.parse_csv()

    def generate_pdf(self):
        """Generar PDF de YAPES"""
        from collections import defaultdict
        
        if not self.data:
            logger.warning("No hay datos para generar PDF")
            return
        
        agrupado = defaultdict(list)
        for item in self.data:
            agrupado[item['nombre_key']].append({
                'nombre': item['nombre'],
                'monto': item['monto'],
                'fecha': item['fecha'],
                'hora': item.get('hora', '')
            })
        
        # Ordenar por nombre normalizado pero guardar nombre original
        nombres = sorted(agrupado.keys())
        
        # Calcular rango de fechas global - usar fechas seleccionadas si existen
        if self.fecha_inicio or self.fecha_fin:
            # Convertir formato YYYY-MM-DD a DD/MM/YYYY para mostrar
            inicio_display = datetime.strptime(self.fecha_inicio, '%Y-%m-%d').strftime('%d/%m/%Y') if self.fecha_inicio else ''
            fin_display = datetime.strptime(self.fecha_fin, '%Y-%m-%d').strftime('%d/%m/%Y') if self.fecha_fin else ''
            if inicio_display and fin_display:
                self.rango_fechas_yape = f"{inicio_display} - {fin_display}"
            elif inicio_display:
                self.rango_fechas_yape = f"{inicio_display} - *"
            else:
                self.rango_fechas_yape = f"* - {fin_display}"
        elif self.data:
            todas_fechas = [item['fecha'] for item in self.data]
            todas_fechas_parsed = sorted(todas_fechas, key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
            self.rango_fechas_yape = f"{todas_fechas_parsed[0]} - {todas_fechas_parsed[-1]}"
        else:
            self.rango_fechas_yape = datetime.now().strftime('%d/%m/%Y')
        
        page_width = 72
        pdf = FPDF(orientation='P', unit='mm', format=(page_width, 300))
        pdf.set_margins(0, 0, 0)
        pdf.set_auto_page_break(auto=True, margin=3)
        pdf.add_page()
        
        pdf.set_font("Arial", 'B', 11)
        pdf.cell(0, 5, "RESUMEN YAPES RECIBIDOS", 0, 1, 'C')
        
        pdf.ln(2)
        pdf.set_font("Arial", '', 9)
        pdf.cell(0, 1, "", "T", 1)
        pdf.ln(2)
        
        for nombre in nombres:
            transacciones = agrupado[nombre]
            
            # Ordenar por fecha y hora ascendente
            transacciones_ordenadas = sorted(
                transacciones,
                key=lambda x: (x['fecha'], x['hora'] if x['hora'] else 'zz')
            )
            
            # Mostrar nombre original (primero encontrado)
            nombre_original = transacciones_ordenadas[0]['nombre'] if transacciones_ordenadas else nombre
            
            # Usar rango del input del usuario
            if self.fecha_inicio and self.fecha_fin:
                inicio_fmt = datetime.strptime(self.fecha_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
                fin_fmt = datetime.strptime(self.fecha_fin, '%Y-%m-%d').strftime('%d/%m/%Y')
                rango = f"{inicio_fmt} - {fin_fmt}"
            else:
                # Calcular de registros reales
                fechas_items = [t['fecha'] for t in transacciones]
                if fechas_items:
                    fechas_parsed = sorted(fechas_items, key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
                    rango = f"{fechas_parsed[0]} - {fechas_parsed[-1]}"
                else:
                    rango = "Sin registros"
            
            # Lista de montos
            count = len(transacciones_ordenadas)

            # PARA:
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 4, f"PARA: {nombre_original}", 0, 1, 'C')
            
            # RANGO DE FECHA:
            pdf.set_font("Arial", '', 12)
            pdf.cell(0, 4, rango, 0, 1, 'C')

            # CANTIDAD REGISTROS:
            pdf.set_font("Arial", '', 10)
            pdf.cell(0, 4, f"N° DE YAPES: {count}", 0, 1, 'C')

            
            pdf.ln(1)
            pdf.set_font("Arial", '', 12)
            pdf.cell(0, 1, "", "T", 1)
            pdf.ln(2)
            
            if count > 26:
                # 3 columnas
                col_width = page_width / 3
                parts = []
                for i in range(3):
                    start = i * ((count + 2) // 3)
                    end = min((i + 1) * ((count + 2) // 3), count)
                    parts.append(transacciones_ordenadas[start:end])
                
                y_start = pdf.get_y()
                heights = [0, 0, 0]
                
                for col_idx, col_items in enumerate(parts):
                    x_pos = col_idx * col_width
                    for txn in col_items:
                        pdf.set_xy(x_pos, y_start + heights[col_idx])
                        pdf.cell(2, 5, "-", 0, 0, 'L')
                        pdf.cell(5, 5, "S/.", 0, 0, 'L')
                        pdf.cell(col_width - 8, 5, f"{txn['monto']:.2f}", 0, 1, 'R')
                        heights[col_idx] += 5
                
                pdf.ln(max(heights) + 2)
                pdf.set_y(y_start + max(heights) + 2)
                
            elif count > 10:
                # 2 columnas
                col_width = page_width / 2
                mid_point = (count + 1) // 2
                
                y_start = pdf.get_y()
                heights = [0, 0]
                
                for txn in transacciones_ordenadas[:mid_point]:
                    pdf.set_xy(0, y_start + heights[0])
                    pdf.cell(2, 5, "-", 0, 0, 'L')
                    pdf.cell(6, 5, "S/.", 0, 0, 'L')
                    pdf.cell(col_width - 10, 5, f"{txn['monto']:.2f}", 0, 1, 'R')
                    heights[0] += 5
                
                for txn in transacciones_ordenadas[mid_point:]:
                    pdf.set_xy(col_width, y_start + heights[1])
                    pdf.cell(2, 5, "-", 0, 0, 'L')
                    pdf.cell(6, 5, "S/.", 0, 0, 'L')
                    pdf.cell(col_width - 8, 5, f"{txn['monto']:.2f}", 0, 1, 'R')
                    heights[1] += 5
                
                pdf.ln(max(heights) + 2)
                pdf.set_y(y_start + max(heights) + 2)
                
            else:
                # 1 columna
                for txn in transacciones_ordenadas:
                    pdf.cell(2, 5, "-", 0, 0, 'L')
                    pdf.cell(6, 5, "S/.", 0, 0, 'L')
                    pdf.cell(0, 5, f"{txn['monto']:.2f}", 0, 1, 'R')
                pdf.ln(3)
            
            # Total
            total_nombre = sum(txn['monto'] for txn in transacciones)
            pdf.set_font("Arial", 'B', 15)
            pdf.cell(0, 5, f"TOTAL: S/. {total_nombre:.2f}", 0, 1, 'R')
            
            pdf.ln(4)
        
        pdf.output(self.output_path)
        logger.info(f"PDF YAPES generado: {self.output_path}")


def log_request(f):
    """Decorador para logging de requests"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        logger.info(f"Request: {request.method} {request.path}")
        return f(*args, **kwargs)
    return decorated_function


def safe_div(a: float, b: float) -> float:
    """División segura que maneja errores de tipo"""
    try:
        return float(a) / float(b)
    except (TypeError, ZeroDivisionError, ValueError):
        return 0.0


class FacturaXMLtoPDF:
    """Clase principal para conversión de XML a PDF"""
    
    def __init__(self, xml_path: str, output_path: str, extra_data: Optional[Dict[str, Any]] = None):
        self.xml_path = xml_path
        self.output_path = output_path
        self.data: Dict[str, Any] = {}
        self.page_width = CONFIG['PAGE_WIDTH']
        self.extra_data = extra_data or {}
        self.errors: list = []
        
    def add_error(self, message: str):
        """Agregar error al historial"""
        self.errors.append(message)
        logger.error(f"Error: {message}")
    
    def parse_xml(self) -> bool:
        """Parsear archivo XML"""
        try:
            if not os.path.exists(self.xml_path):
                self.add_error(f"Archivo no encontrado: {self.xml_path}")
                return False
            
            tree = ET.parse(self.xml_path)
            root = tree.getroot()
            
            namespaces = {
                'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
                'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
                'ds': 'http://www.w3.org/2000/09/xmldsig#'
            }
            
            # Datos básicos
            self.data['monto_letras'] = ''
            self.data['forma_pago'] = ''
            
            # Extraer notes
            for note in root.findall('.//cbc:Note', namespaces):
                note_text = note.text or ''
                if note.get('languageLocaleID') == "1000":
                    self.data['monto_letras'] = note_text
                elif note.get('languageID') == "L":
                    self.data['forma_pago'] = note_text
            
            # Datos principales
            self.data['numero_factura'] = self._get_text(root, './/cbc:ID', namespaces)
            self.data['fecha_emision'] = self._get_text(root, './/cbc:IssueDate', namespaces)
            self.data['hora_emision'] = self._get_text(root, './/cbc:IssueTime', namespaces)
            
            # Tipo de documento
            numero = self.data.get('numero_factura', '')
            self.data['tipo_documento'] = "FACTURA" if numero and numero[0].upper() == 'F' else "BOLETA DE VENTA"
            
            # Datos del emisor
            emisor = root.find('.//cac:AccountingSupplierParty/cac:Party', namespaces)
            if emisor is not None:
                self.data['emisor_nombre'] = self._get_text(emisor, './/cbc:Name', namespaces) or \
                                             self._get_text(emisor, './/cbc:RegistrationName', namespaces)
                self.data['emisor_ruc'] = self._get_text(emisor, './/cac:PartyIdentification/cbc:ID', namespaces) or \
                                          self._get_text(emisor, './/cbc:ID', namespaces)
                self.data['emisor_direccion'] = self._get_text(emisor, './/cac:AddressLine/cbc:Line', namespaces)
                self.data['emisor_distrito'] = self._get_text(emisor, './/cbc:District', namespaces)
                self.data['emisor_departamento'] = self._get_text(emisor, './/cbc:CityName', namespaces)
            
            # Datos del cliente
            cliente = root.find('.//cac:AccountingCustomerParty/cac:Party', namespaces)
            if cliente is not None:
                self.data['cliente_nombre'] = self._get_text(cliente, './/cbc:RegistrationName', namespaces)
                self.data['cliente_ID'] = self._get_text(cliente, './/cac:PartyIdentification/cbc:ID', namespaces) or \
                                         self._get_text(cliente, './/cbc:ID', namespaces)
                self.data['cliente_direccion'] = self._get_text(cliente, './/cac:AddressLine/cbc:Line', namespaces)
                self.data['cliente_distrito'] = self._get_text(cliente, './/cbc:District', namespaces)
                self.data['cliente_provincia'] = self._get_text(cliente, './/cbc:CountrySubentity', namespaces)
                self.data['cliente_departamento'] = self._get_text(cliente, './/cbc:CityName', namespaces)
            
            # Guía de remisión
            self.data['cliente_guia'] = self._get_text(root, './/cac:DespatchDocumentReference/cbc:ID', namespaces)
            
            # Totales
            self.data['total_venta'] = float(self._get_text(root, './/cac:TaxSubtotal/cbc:TaxableAmount', namespaces, '0.0'))
            self.data['total_igv'] = float(self._get_text(root, './/cac:TaxTotal/cbc:TaxAmount', namespaces, '0.0'))
            self.data['total_pagar'] = float(self._get_text(root, './/cac:LegalMonetaryTotal/cbc:PayableAmount', namespaces, '0.0'))
            
            # Items
            self.data['items'] = []
            for item in root.findall('.//cac:InvoiceLine', namespaces):
                item_data = {
                    'id': self._get_text(item, './/cac:SellersItemIdentification/cbc:ID', namespaces),
                    'unidad': self._get_text(item, './/cbc:Note', namespaces),
                    'descripcion': self._get_text(item, './/cbc:Description', namespaces),
                    'cantidad': float(self._get_text(item, './/cbc:InvoicedQuantity', namespaces, '0')),
                    'precio_unitario': float(self._get_text(item, './/cac:Price/cbc:PriceAmount', namespaces, '0.00')),
                    'total': float(self._get_text(item, './/cbc:LineExtensionAmount', namespaces, '0.00'))
                }
                self.data['items'].append(item_data)
            
            # Datos para QR
            self.data['tipo_codigo_invoice'] = self._get_text(root, './/cbc:InvoiceTypeCode', namespaces)
            self.data['tipo_doc_cli'] = self._get_text(
                root, 
                './/cac:AccountingCustomerParty/cac:Party/cac:PartyIdentification/cbc:ID', 
                namespaces, 
                attr='schemeID'
            )
            
            # Digest value
            namespaces_ds = {'ds': 'http://www.w3.org/2000/09/xmldsig#'}
            digest_node = root.find('.//ds:DigestValue', namespaces_ds)
            self.data['digest_value'] = digest_node.text if digest_node is not None else ""
            
            self.data.update(self.extra_data)
            
            logger.info(f"XML parseado: {self.data.get('numero_factura', 'N/A')}")
            return True
            
        except Exception as e:
            self.add_error(f"Error al parsear XML: {str(e)}")
            return False
    
    def _get_text(self, node, path: str, namespaces: Dict, default: str = '-', attr: str = None) -> str:
        """Obtener texto de manera segura"""
        if node is None:
            return default
        try:
            found = node.find(path, namespaces)
            if found is not None:
                if attr:
                    return found.get(attr, default)
                return found.text if found.text else default
        except Exception:
            pass
        return default
    
    def format_currency(self, amount: str) -> str:
        """Formatear monto como moneda"""
        try:
            return f"S/. {float(amount):.2f}"
        except (ValueError, TypeError):
            return "S/. 0.00"
    
    def calculate_total_height(self) -> float:
        """Calcular altura total del PDF"""
        pdf_temp = FPDF(orientation='P', unit='mm', format=(self.page_width, 300))
        pdf_temp.set_margins(0, 0, 0)
        pdf_temp.add_page()
        pdf_temp.set_font("Arial", '', 10)
        
        y = 2
        
        # Logo
        if os.path.exists("images/logo_manchester.png"):
            y += safe_div(40, 3) + 6
        
        # Encabezado emisor
        y += 4
        emisor_nombre = self.data.get('emisor_nombre', 'N/A')
        y += 8 if len(emisor_nombre) > 35 else 4
        y += 4
        y += 4  # Dirección
        y += 1 + 2 + 4 + 5 + 1 + 2 + 4 + 4 + 1 + 2 + 4 + 1 + 2
        
        # Cliente
        cliente_nombre = self.data.get('cliente_nombre', 'N/A').upper()
        y += 8 if len(cliente_nombre) > 25 else 4
        y += 1
        c_dir = (self.data.get('cliente_direccion', '') or "").strip()
        y += 8 if len(c_dir) > 24 else 4
        
        extras = [p for p in [
            self.data.get('cliente_departamento', ''),
            self.data.get('cliente_provincia', ''),
            self.data.get('cliente_distrito', '')
        ] if p and p.strip()]
        y += 1 + (8 if len(" - ".join(extras)) > 24 else 4) if extras else 0
        
        y += 2 + 1 + 2 + 4 + 1 + 1 + 2 + 1 + 2 + 1 + 2
        
        # Items
        pdf_temp.set_font("Arial", '', 8)
        for item in self.data.get('items', []):
            descripcion = str(item.get('descripcion', ''))
            lineas = 1
            if descripcion:
                palabras = descripcion.split()
                linea_actual = ""
                for palabra in palabras:
                    prueba_linea = f"{linea_actual} {palabra}".strip()
                    if pdf_temp.get_string_width(prueba_linea) <= 22:
                        linea_actual = prueba_linea
                    else:
                        lineas += 1
                        linea_actual = palabra
            y += max(lineas, 1) * 3 + 2
        
        y += 2 + 5 + 5 + 8 + 2
        
        # Totales
        monto_l = self.data.get('monto_letras', '')
        y += 8 if len(monto_l) > 24 else 4
        
        y += safe_div(30, 3) + 20 + 4 + 4 + 20
        
        return max(100, min(800, y + 30))
    
    def generate_pdf(self, output_format: str = 'ticket'):
        """Generar PDF según formato"""
        if output_format == 'ticket':
            self._generate_ticket_pdf()
        elif output_format == 'shipping_label':
            self._generate_shipping_label_pdf()
        else:
            raise ValueError(f"Formato no válido: {output_format}")
    
    def _generate_shipping_label_pdf(self):
        """Generar etiqueta de envío 100x150mm sin márgenes"""
        label_width = 100
        label_height = 150
        
        pdf = FPDF(orientation='P', unit='mm', format=(label_width, label_height))
        pdf.set_margins(0, 0, 0)

        pdf.add_page()
        pdf.set_auto_page_break(auto=False)

        
        emisor_nombre = self.data.get('emisor_nombre', 'N/A').upper()
        emisor_ruc = self.data.get('emisor_ruc', 'N/A')
        
        logo_path = "images/logo_manchester.png"
        logo_width = 30
        pdf.set_y(0)
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=8, y=2, w=logo_width)
            pdf.set_y(20)
        else:
            pdf.set_font("Arial", 'B', 14)
            pdf.set_xy(1, 1)
            pdf.cell(35, 7, "MANCHESTER", 0, 0, 'L')
            pdf.set_y(12)
        
        # REMITENTE pegado al borde derecho
        remitente_x = label_width - 50.2
        pdf.set_y(0)
        pdf.set_draw_color(0, 0, 0)
        pdf.set_line_width(0.3)
        
        pdf.set_font("Arial", 'B', 9)
        pdf.set_xy(remitente_x, 0)
        pdf.cell(50, 5, "REMITENTE", 'TLR', 1, 'C')
        
        pdf.set_font("Arial", '', 9)
        pdf.set_xy(remitente_x, pdf.get_y())
        
        pdf.cell(50, 4, f"RUC: {emisor_ruc}", 'LR', 1, 'C')
        
        
        pdf.set_font("Arial", 'B', 9)
        pdf.set_xy(remitente_x, pdf.get_y())
        
        if pdf.get_string_width(emisor_nombre) > 40:
            pdf.set_font("Arial", 'B', 8)
            pdf.multi_cell(50, 4, emisor_nombre, 'BLR', 'C')
        else:
            pdf.cell(50, 4, emisor_nombre, 'BLR', 1, 'C')
        
        # Línea divisoria pegada al borde
        pdf.line(0, pdf.get_y(), label_width, pdf.get_y())
        
        contenido_destinatario = label_width - 5

        # DATOS DESTINATARIO pegados al borde izquierdo
        # Si "recoje otra persona" está activo, usar esos datos
        recoje_otra_persona = self.data.get('recoje_otra_persona', False)
        
        if recoje_otra_persona:
            cliente_id = self.data.get('recoje_dni', '')
            cliente_nombre = self.data.get('recoje_nombre', 'N/A').upper()
            cliente_dir = self.data.get('recoje_direccion', '')
            cliente_dis = ''
            cliente_pro = ''
            cliente_dep = ''
        else:
            cliente_id = self.data.get('cliente_ID', '')
            cliente_nombre = self.data.get('cliente_nombre', 'N/A').upper()
            cliente_dir = self.data.get('cliente_direccion', '')
            cliente_dis = self.data.get('cliente_distrito', '')
            cliente_pro = self.data.get('cliente_provincia', '')
            cliente_dep = self.data.get('cliente_departamento', '')

        pdf.set_margins(2, 0, 2)

        pdf.ln(1)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 5, "DATOS DESTINATARIO", 0, 1, 'C')

        pdf.set_font("Arial", 'B', 26)
        
        if pdf.get_string_width(cliente_nombre) > 190:
            pdf.set_font("Arial", 'B', 22)
            pdf.multi_cell(contenido_destinatario, 7, cliente_nombre, 0, 'L')
        else:
            pdf.multi_cell(contenido_destinatario, 10, cliente_nombre, 0, 'L')
        
        pdf.ln(2)
        
        pdf.set_font("Arial", '', 18)
        
        id_label = "RUC" if len(cliente_id) == 11 else ("DNI" if len(cliente_id) == 8 else "CE")
        pdf.cell(contenido_destinatario, 5, f"{id_label}: {cliente_id}", 0, 1, 'L')
        
        pdf.ln(1)
        
        invalidos = ['N/A', 'n/a', '-', '--', '---', '', None]
        
        c_dir = cliente_dir.strip() if cliente_dir and cliente_dir.strip() not in invalidos else "-"
        
        direccion_partes = [cliente_dep, cliente_pro, cliente_dis]
        direccion_ciudad = [p.strip() for p in direccion_partes if p and p.strip() not in invalidos]
        direccion_completa = " - ".join(direccion_ciudad).upper()

        tiene_ciudad = any(p and p.strip() not in invalidos for p in direccion_partes)

        pdf.set_font("Arial", '', 8)
        pdf.cell(contenido_destinatario, 5, "DIRECCIÓN DE ENVIO:", 0, 1, 'L')
        pdf.set_font("Arial", 'B', 18)
        pdf.ln(1)
        
        pdf.multi_cell(contenido_destinatario, 7, c_dir.upper(), 'B', 'L')
        
        if tiene_ciudad:
            pdf.ln(1)
            pdf.set_font("Arial", '', 8)
            pdf.cell(contenido_destinatario, 5, "CIUDAD:", 0, 1, 'L')
            pdf.set_font("Arial", 'B', 15)
            pdf.ln(1)
            pdf.multi_cell(contenido_destinatario, 7, direccion_completa, 'B', 'L')
        
        pdf.line(0, pdf.get_y(), label_width, pdf.get_y())

        
        pdf.ln(3)
        
        agency_name = self.data.get('agency_name', '').upper()
        other_notes = self.data.get('other_notes', '').upper()
        
        if agency_name:
            ancho_label = 22
            ancho_valor = label_width - ancho_label
            
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(ancho_label, 6, "AGENCIA:", 0, 0, 'L')
            
            if pdf.get_string_width(agency_name) > ancho_valor:
                pdf.set_font("Arial", 'B', 12)
                pdf.multi_cell(ancho_valor, 6, agency_name, 0, 'L')
            else:
                pdf.multi_cell(ancho_valor, 6, agency_name, 0, 'L')
            pdf.ln(1)
        
        if other_notes:
            ancho_label = 18
            ancho_valor = 72
            
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(ancho_label, 6, "OTROS:", 0, 0, 'L')
            
            pdf.set_font("Arial", '', 12)
            
            if pdf.get_string_width(other_notes) > 100:
                pdf.set_font("Arial", '', 12)
                pdf.multi_cell(ancho_valor, 6, other_notes, 0, 'L')
            else:
                pdf.multi_cell(ancho_valor, 6, other_notes, 0, 'L')
            pdf.ln(1)
        
        # QR en extremo inferior izquierdo
        qr_width = 20
        qr_x = 1
        qr_y = label_height - qr_width - 5

        pdf.set_font("Arial", 'B', 10)
        pdf.set_xy(0, qr_y - 5)
        pdf.cell(100, 4, "WWW.MANCHESTERCOLLECTIONPERU.COM", 0, 0, 'C')
        
        qr_path = "images/qr_mostrario.png"
        if os.path.exists(qr_path):
            pdf.image(qr_path, x=qr_x, y=qr_y, w=qr_width)
        else:
            pdf.set_fill_color(200, 200, 200)
            pdf.rect(qr_x, qr_y, qr_width, qr_width, 'F')
            pdf.set_font("Arial", '', 8)
            pdf.set_xy(qr_x, qr_y + 12)
            pdf.cell(qr_width, 5, "QR", 0, 0, 'C')
        
        pdf.set_font("Arial", 'B', 8)
        pdf.set_xy(qr_x, qr_y + qr_width)
        pdf.cell(qr_width, 4, "CATÁLOGO", 0, 0, 'C')

        num_documento = self.data.get('numero_factura', 'N/A')
        fecha = self.data.get('fecha_emision', 'N/A')
        guia_remision = self.data.get('cliente_guia', 'N/A')
        
        fecha_formateada = fecha
        
        if fecha != 'N/A':
            try:
                objeto_fecha = datetime.strptime(fecha, '%Y-%m-%d')
                fecha_formateada = objeto_fecha.strftime('%d/%m/%Y')
            except ValueError:
                pass
        
        # Info inmediatamente después del QR (misma altura Y)
        info_x = qr_x + qr_width + 1
        info_width = label_width - info_x - 1
        
        pdf.set_xy(info_x, qr_y +0.5)
        pdf.set_draw_color(0, 0, 0)
        pdf.set_line_width(0.3)
        pdf.set_fill_color(255, 255, 255)
        
        pdf.set_font("Arial", 'B', 12)
        
        pdf.cell(info_width, 7.5, f"N° DE DOC.: {num_documento}", 1, 1, 'C', True)
        
        if guia_remision.strip() not in ['', 'N/A', '-', 'None']:
            pdf.set_x(info_x)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(info_width, 7.5, f"GUÍA DE REMISIÓN: N° {guia_remision}", 1, 1, 'C', True)
        
        pdf.set_x(info_x)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(info_width, 7.5, f"FECHA DE EMISIÓN: {fecha_formateada}", 1, 1, 'C', True)
        
        
        pdf.output(self.output_path)
        logger.info(f"Etiqueta de envío generada: {self.output_path} (100mm x 150mm)")

    def _generate_qr(self, pdf):
        """Generar código QR"""
        ruc = self.data.get('emisor_ruc', '')
        tipo = self.data.get('tipo_codigo_invoice', '01')
        serie, correlativo = self.data.get('numero_factura', '001-000001').split('-')[:2]
        
        val_igv = safe_div(float(self.data.get('total_igv', '0')), 1)
        val_total = safe_div(float(self.data.get('total_pagar', '0')), 1)
        
        fecha = self.data.get('fecha_emision', '')
        tipo_doc_cli = self.data.get('tipo_doc_cli', '')
        ruc_cli = self.data.get('cliente_ID', '')
        digest = self.data.get('digest_value', '')
        
        cadena_qr = f"{ruc}|{tipo}|{serie}|{correlativo}|{val_igv:.2f}|{val_total:.2f}|{fecha}|{tipo_doc_cli}|{ruc_cli}|{digest}|"
        
        try:
            qr_img = qrcode.make(cadena_qr)
            qr_path = "images/temp_qr.png"
            qr_img.save(qr_path)
            
            if ruc and os.path.exists(qr_path):
                img_w = 30
                img_x = safe_div(self.page_width - img_w, 2)
                pdf.image(qr_path, x=img_x, y=pdf.get_y(), w=img_w)
                pdf.set_y(pdf.get_y() + safe_div(img_w, 3) + 20)
        except Exception as e:
            logger.warning(f"Error generando QR: {e}")        
    
    def _generate_ticket_pdf(self):
        """Generar ticket 72mm"""

        page_height = self.calculate_total_height()
        
        pdf = FPDF(orientation='P', unit='mm', format=(self.page_width, page_height))
        pdf.set_margins(0, 0, 0)
        pdf.set_auto_page_break(auto=False)
        pdf.add_page()
        
        # Logo
        image_path = "images/logo_manchester.png"
        if os.path.exists(image_path):
            try:
                pdf.image(image_path, x=(self.page_width - 40) / 2, y=3, w=40)
                pdf.ln(safe_div(40, 3) + 4)
            except Exception as e:
                logger.warning(f"Error cargando logo: {e}")
                pdf.ln(5)
        else:
            pdf.ln(5)
        
        pdf.ln(1)
        
        # Encabezado emisor
        pdf.set_font("Arial", '', 10)
        emisor_nombre = self.data.get('emisor_nombre', 'N/A')
        if len(emisor_nombre) > 35:
            pdf.multi_cell(0, 4, emisor_nombre, 0, 'C')
        else:
            pdf.cell(0, 4, emisor_nombre, 0, 1, 'C')
        
        pdf.set_font("Arial", '', 10)
        pdf.cell(0, 4, f"RUC: {self.data.get('emisor_ruc', 'N/A')}", 0, 1, 'C')
        
        # Dirección emisor
        emisor_dir = self.data.get('emisor_direccion', '')
        emisor_dis = self.data.get('emisor_distrito', '')
        emisor_dep = self.data.get('emisor_departamento', '')
        dir_completa = f"{emisor_dir}"
        if emisor_dis:
            dir_completa += f" - {emisor_dis}"
        if emisor_dep:
            dir_completa += f" - {emisor_dep}"
        
        if len(dir_completa) > 35:
            pdf.multi_cell(0, 4, dir_completa, 0, 'C')
        else:
            pdf.cell(0, 4, dir_completa, 0, 1, 'C')
        
        pdf.ln(1)
        pdf.cell(0, 1, "", "T", 1)
        pdf.ln(2)
        
        # Tipo documento
        pdf.set_font("Arial", '', 10)
        pdf.cell(0, 4, f"{self.data.get('tipo_documento', 'COMPROBANTE')} ELECTRÓNICA", 0, 1, 'C')
        pdf.set_font("Arial", 'B', 14)
        
        pdf.cell(0, 5, self.data.get('numero_factura', 'N/A'), 0, 1, 'C')
        
        
        pdf.ln(2)
        pdf.cell(0, 1, "", "T", 1)
        pdf.ln(2)
        
        # Datos cliente
        pdf.set_font("Arial", '', 10)
        cliente_id = self.data.get('cliente_ID', '')
        label_id = "RUC:" if len(cliente_id) == 11 else ("DNI:" if len(cliente_id) == 8 else "CE:")
        label_w = pdf.get_string_width(label_id) + 2
        pdf.cell(label_w, 4, label_id, 0, 0)
        pdf.set_font("Arial", '', 12)
        pdf.cell(self.page_width - label_w, 4, cliente_id, 0, 1)
        
        pdf.ln(1)
        pdf.set_font("Arial", '', 10)
        cliente_nombre = self.data.get('cliente_nombre', 'N/A').upper()
        label = "RAZÓN SOCIAL:" if len(cliente_id) == 11 else "CLIENTE:"
        if len(cliente_nombre) > 20:
            pdf.multi_cell(0, 4, f"{label} {cliente_nombre}", 0, 'L')
        else:
            pdf.cell(0, 4, f"{label} {cliente_nombre}", 0, 1, 'L')
        
        pdf.ln(1)
        
        # Dirección cliente
        c_dir = (self.data.get('cliente_direccion', '') or "").strip()
        c_dis = self.data.get('cliente_distrito', '')
        c_pro = self.data.get('cliente_provincia', '')
        c_dep = self.data.get('cliente_departamento', '')
        
        pdf.set_font("Arial", '', 10)
        if len(c_dir) > 22:
            pdf.multi_cell(0, 4, f"DIRECCIÓN: {c_dir.upper()}", 0, 'L')
        else:
            pdf.cell(0, 4, f"DIRECCIÓN: {c_dir.upper()}", 0, 1, 'L')
        
        extras = [p for p in [c_dep, c_pro, c_dis] if p and p.strip()]
        if extras:
            pdf.ln(1)
            txt = f"CIUDAD: {' - '.join(extras).upper()}"
            if len(txt) > 24:
                pdf.multi_cell(0, 4, txt, 0, 'L')
            else:
                pdf.cell(0, 4, txt, 0, 1, 'L')
        
        pdf.ln(2)
        pdf.cell(0, 1, "", "T", 1)
        pdf.ln(2)
        
        # Guía de remisión
        guia = self.data.get('cliente_guia', '')
        if guia and guia not in ['', 'N/A', '-']:
            pdf.cell(35, 4, "GUÍA DE REMISIÓN: ", 0, 0)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(37, 4, f"N° {guia}", 0, 1)
            pdf.ln(1)
        
        # Forma de pago
        pdf.set_font("Arial", '', 10)
        f_pago = self.data.get('forma_pago', '').upper()
        pdf.cell(0, 4, f"FORMA DE PAGO: {f_pago}" if f_pago else "", 0, 1)
        
        pdf.ln(1)
        # Fecha
        fecha = self.data.get('fecha_emision', 'N/A')
        if fecha != 'N/A':
            try:
                fecha = datetime.strptime(fecha, '%Y-%m-%d').strftime('%d/%m/%Y')
            except ValueError:
                pass
        hora = self.data.get('hora_emision', '')
        pdf.cell(0, 4, f"FECHA DE EMISIÓN: {fecha} {hora}".strip(), 0, 1)
        
        pdf.ln(2)
        pdf.cell(0, 1, "", "T", 1)
        pdf.ln(2)
        
        # Tabla de items
        anchuras = [7, 13, 9, 20, 9, 14]
        pdf.set_draw_color(255, 255, 255)
        pdf.set_font("Arial", '', 7)
        
        headers = ["COD.", "CANT.", "UNID.", "DESCRIPCIÓN", "V.UNIT.", "V.VENTA"]

        for i, h in enumerate(headers):
            pdf.cell(anchuras[i], 1, h, 1, 0, 'C')
        
        pdf.set_draw_color(0, 0, 0)

        pdf.ln(2)
        pdf.cell(0, 2, "", "T", 1)
        
        
        pdf.set_font("Arial", '', 8)
        
        pdf.set_draw_color(255, 255, 255)
        
        pdf.set_font("Arial", '', 8)
        for item in self.data.get('items', []):
            x_start = pdf.get_x()
            y_start = pdf.get_y()

            pdf.cell(anchuras[0], 3, str(item.get('id', ''))[:20], 1, 0, 'C')
            pdf.set_font("Arial", '', 10)

            
            cantidad = float(item.get('cantidad', '0'))

            pdf.cell(anchuras[1], 3, f"{cantidad:.2f}"[:6], 1, 0, 'C')
            
            pdf.set_font("Arial", '', 8)

            pdf.cell(anchuras[2], 3, str(item.get('unidad', 'MTS'))[:4], 1, 0, 'C')
            x_descripcion = pdf.get_x()


            pdf.set_font("Arial", '', 7)
            pdf.multi_cell(anchuras[3], 3, str(item.get('descripcion', '')), 0, 'C')
            y_final = pdf.get_y()
            
            altura_fila = y_final - y_start

            pdf.set_y(y_start)
            pdf.set_font("Arial", '', 8)
            
            pdf.set_x(x_descripcion + anchuras[3]) 
            pdf.cell(anchuras[4], altura_fila, str(item.get('precio_unitario', ''))[:5], 1, 0, 'C')
            pdf.cell(anchuras[5], altura_fila, str(item.get('total', '')), 1, 1, 'C')
            pdf.set_xy(x_start, pdf.get_y())
            pdf.ln(2)
        
        pdf.ln(2)
        
        # Totales
        pdf.set_font("Arial", '', 10)
        pdf.cell(45, 5, "OP. GRAVADA:", 0, 0)
        pdf.cell(27, 5, self.format_currency(self.data.get('total_venta', '0.00')), 0, 1, 'R')
        pdf.cell(45, 5, "IGV:", 0, 0)
        pdf.cell(27, 5, self.format_currency(self.data.get('total_igv', '0.00')), 0, 1, 'R')
        
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(45, 6, "TOTAL:", 0, 0)
        pdf.cell(27, 8, self.format_currency(self.data.get('total_pagar', '0.00')), 0, 1, 'R')
        
        pdf.ln(2)
        pdf.set_font("Arial", '', 10)
        monto_l = self.data.get('monto_letras', '')
        if monto_l:
            pdf.multi_cell(0, 4, f"SON: {monto_l}", 0) if len(monto_l) > 24 else \
                pdf.cell(0, 4, f"SON: {monto_l}", 0, 1)
        
        pdf.ln(2)
        
        # QR
        self._generate_qr(pdf)
        
        # Pie
        pdf.set_font("Arial", '', 9)

        pdf.cell(0, 4, "Representación impresa del comprobante de pago", 0,1, 'C')
        pdf.set_font("Arial", 'I', 10)
        pdf.cell(0, 4, "¡Gracias por su compra!", 0, 1, 'C')
        
        pdf.output(self.output_path)
        logger.info(f"PDF generado: {self.output_path}")
    



# ========== FLASK ROUTES ==========

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Conversor XML a PDF</title>
    <link rel="icon" href="/images/favicon.ico?v=2" type="image/x-icon">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/sweetalert2@11/dist/sweetalert2.min.css">
    <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
    <script>
    function downloadPDF(base64Data, filename) {
        var link = document.createElement('a');
        link.href = 'data:application/pdf;base64,' + base64Data;
        link.download = filename;
        link.click();
    }
    </script>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', Arial, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container { max-width: 1400px; margin: 0 auto; }
        h1 { 
            color: white; 
            text-align: center; 
            margin-bottom: 10px;
            font-size: 2rem;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }
        .subtitle {
            color: rgba(255,255,255,0.9);
            text-align: center;
            margin-bottom: 30px;
            font-size: 1rem;
        }
        .main-content {
            display: flex;
            gap: 20px;
            align-items: flex-start;
        }
        .left-panel {
            flex: 1;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            padding: 20px;
            min-height: 600px;
        }
        .right-panel {
            width: 380px;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            padding: 25px;
        }
        .viewer-header {
            display: flex;
            gap: 10px;
            margin-bottom: 15px;
        }
        .viewer-header .btn {
            flex: 1;
            padding: 10px;
            font-size: 0.85rem;
        }
        @media (max-width: 1024px) {
            .main-content {
                flex-direction: column-reverse;
            }
            .right-panel {
                width: 100%;
            }
            .left-panel {
                min-height: 500px;
            }
        }
        @media (max-width: 600px) {
            .container {
                padding: 10px;
            }
            h1 {
                font-size: 1.5rem;
            }
            .subtitle {
                font-size: 0.9rem;
            }
            .right-panel, .left-panel {
                padding: 15px;
                min-height: auto;
            }
            #pdf-viewer {
                height: 400px;
            }
            .viewer-header {
                flex-direction: column;
            }
            .viewer-header .btn {
                width: 100%;
            }
            .info p {
                font-size: 0.8rem;
            }
            .btn-group {
                flex-direction: column;
            }
        }
        .form-group { margin-bottom: 20px; }
        label { 
            display: block; 
            margin-bottom: 8px; 
            font-weight: 600; 
            color: #374151;
            font-size: 0.95rem;
        }
        .file-input-wrapper {
            position: relative;
        }
        .file-name {
            display: block;
            padding: 12px 15px;
            background: #f0fdf4;
            border: 2px solid #22c55e;
            border-radius: 10px;
            color: #166534;
            font-weight: 500;
            margin-bottom: 10px;
        }
        .file-name.empty {
            background: #f8fafc;
            border: 2px dashed #cbd5e1;
            color: #9ca3af;
        }
        .file-input {
            display: block;
            width: 100%;
            padding: 15px;
            border: 2px dashed #cbd5e1;
            border-radius: 10px;
            background: #f8fafc;
            cursor: pointer;
            transition: all 0.3s;
            text-align: center;
            color: #64748b;
            font-size: 0.9rem;
        }
        .file-input:hover {
            border-color: #667eea;
            background: #f1f5f9;
        }
        .file-input.has-file {
            border-color: #22c55e;
            background: #f0fdf4;
            color: #166534;
        }
        input[type="file"] {
            display: none;
        }
        .file-label {
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 60px;
            border: 3px dashed #cbd5e1;
            border-radius: 15px;
            background: #f8fafc;
            cursor: pointer;
            transition: all 0.3s;
            text-align: center;
            color: #64748b;
            font-size: 1.3rem;
            font-weight: bold;
            min-height: 150px;
        }
        .file-label:hover {
            border-color: #667eea;
            background: #f1f5f9;
        }
        .file-label.active {
            border-color: #22c55e;
            background: #f0fdf4;
            color: #166534;
            padding: 15px;
            min-height: auto;
            font-size: 1rem;
            justify-content: space-between;
        }
        .file-container {
            display: flex;
            align-items: center;
            gap: 10px;
            margin-top: 10px;
        }
        .file-container .file-name {
            flex: 1;
            padding: 12px 15px;
            background: #f0fdf4;
            border: 2px solid #22c55e;
            border-radius: 10px;
            color: #166534;
            font-weight: 500;
            word-break: break-all;
        }
        .file-x-btn {
            padding: 8px 12px;
            background: #ef4444;
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 14px;
            font-weight: bold;
        }
        .file-x-btn:hover {
            background: #dc2626;
        }
        select { 
            width: 100%; 
            padding: 15px; 
            border-radius: 10px; 
            border: 1px solid #cbd5e1;
            background: white;
            font-size: 1rem;
        }
        textarea {
            width: 100%;
            padding: 15px;
            border-radius: 10px;
            border: 1px solid #cbd5e1;
            background: white;
            font-size: 1rem;
            font-family: inherit;
            resize: vertical;
        }
        textarea:focus, select:focus, input:focus {
            outline: none;
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }
        input[type="text"], input[type="number"] {
            width: 100%;
            padding: 12px 15px;
            border-radius: 10px;
            border: 1px solid #cbd5e1;
            background: white;
            font-size: 1rem;
            font-family: inherit;
            box-sizing: border-box;
        }
        input[type="checkbox"] {
            width: 18px;
            height: 18px;
            cursor: pointer;
        }
        .btn { 
            width: 100%; 
            padding: 15px; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white; 
            border: none; 
            border-radius: 10px; 
            font-size: 1rem; 
            font-weight: 600; 
            cursor: pointer;
            transition: transform 0.2s, box-shadow 0.2s;
        }
        .btn:hover { 
            transform: translateY(-2px);
            box-shadow: 0 10px 20px rgba(102, 126, 234, 0.4);
        }
        .btn:disabled {
            background: #9ca3af;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }
        .btn-convert { margin-top: 10px; }
        .btn-download { 
            background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%);
        }
        .btn-clean { 
            background: linear-gradient(135deg, #b91c1c 0%, #991b1b 100%);
            text-decoration: none;
        }
        .btn-group {
            display: flex;
            gap: 10px;
            margin-top: 10px;
        }
        .btn-group .btn {
            flex: 1;
            padding: 12px;
            font-size: 0.9rem;
        }
        #pdf-viewer {
            width: 100%;
            height: 550px;
            border-radius: 8px;
            overflow: hidden;
        }
        .info { 
            margin-top: 20px; 
            padding: 20px; 
            background: #f0fdf4; 
            border-radius: 10px;
            border-left: 4px solid #22c55e;
        }
        .info h3 { color: #16a34a; margin-bottom: 12px; font-size: 1rem; }
        .info p { color: #166534; margin-bottom: 6px; line-height: 1.5; font-size: 0.9rem; }
        .info strong { display: inline-block; width: 80px; }
        .error { 
            background: #fef2f2; 
            color: #dc2626; 
            padding: 15px; 
            border-radius: 10px;
            border-left: 4px solid #dc2626;
            margin-top: 15px;
            font-size: 0.9rem;
        }
        .empty-state {
            text-align: center;
            color: #9ca3af;
            padding: 120px 20px;
        }
        .empty-state h2 { margin-bottom: 10px; color: #6b7280; }
        .footer {
            text-align: center;
            color: rgba(255,255,255,0.7);
            margin-top: 30px;
            font-size: 0.85rem;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Conversor XML a PDF</h1>
        
        <div class="main-content">
            <div class="right-panel">
                <form id="convertirForm" method="POST" action="/convertir" enctype="multipart/form-data" onsubmit="return validarArchivo()">
                    <div class="form-group">
                        <label>Seleccionar archivo XML:</label>
                        <label for="xml_file" class="file-label" id="fileLabel">
                            Subir archivo
                        </label>
                        <input type="file" name="xml_file" id="xml_file" accept=".xml,.csv,.xlsx" onchange="updateFileName()">
                    </div>
                    
                    <div class="form-group">
                        <label for="formato">Formato de salida:</label>
                        <select name="formato" id="formato" onchange="checkFormato(); toggleAgencia()">
                            <option value="ticket" {% if selected_formato=='ticket' %}selected{% endif %}>Ticket 72mm</option>
                            <option value="shipping_label" {% if selected_formato=='shipping_label' %}selected{% endif %}>Etiqueta de Envío 100mmx150mm</option>
                            <option value="yapes" {% if selected_formato=='yapes' %}selected{% endif %}>Yapes Resumen 72mm</option>
                        </select>
                    </div>
                    
                    <div class="form-group" id="recojeGroup" style="display: none;">
                        <label style="display: flex; align-items: center; gap: 8px; cursor: pointer;">
                            <input type="checkbox" name="recoje_otra_persona" id="recoje_otra_persona" onchange="toggleRecojeOtraPersona()" style="width: 18px; height: 18px;"
                            {% if selected_recoje %}checked{% endif %}>
                            <span style="font-weight: bold;">Recoge otra persona</span>
                        </label>
                    </div>
                    
                    <div class="form-group" id="yapesFechaGroup" style="display: none;">
                        <label style="font-weight: bold; margin-bottom: 8px; display: block;">Rango de Fechas:</label>
                        <div style="display: flex; gap: 10px; align-items: center; flex-wrap: wrap;">
                            <div style="flex: 1; min-width: 120px;">
                                <input type="date" name="yapes_fecha_inicio" id="yapes_fecha_inicio" 
                                    value="{{ selected_yapes_fecha_inicio or '' }}"
                                    style="width: 100%; padding: 10px; border: 2px solid #e2e8f0; border-radius: 8px; font-size: 14px; background: white;">
                            </div>
                            <span style="font-weight: bold; color: #667eea;">-</span>
                            <div style="flex: 1; min-width: 120px;">
                                <input type="date" name="yapes_fecha_fin" id="yapes_fecha_fin" 
                                    value="{{ selected_yapes_fecha_fin or '' }}"
                                    style="width: 100%; padding: 10px; border: 2px solid #e2e8f0; border-radius: 8px; font-size: 14px; background: white;">
                            </div>
                        </div>
                        <small style="color: #718096; font-size: 12px; margin-top: 4px; display: block;">Seleccione el rango de fechas para filtrar los registros YAPES</small>
                    </div>
                    
                    <div class="form-group" id="recojeDatosGroup" style="display: none;">
                        <label for="recoje_dni">DNI:</label>
                        <input type="text" name="recoje_dni" id="recoje_dni" maxlength="8" pattern="[0-9]{8}" value="{{ selected_recoje_dni or '' }}">
                        
                        <label for="recoje_nombre">Nombres y Apellidos:</label>
                        <input type="text" name="recoje_nombre" id="recoje_nombre" maxlength="80" value="{{ selected_recoje_nombre or '' }}">
                        
                        <label for="recoje_direccion">Dirección:</label>
                        <input type="text" name="recoje_direccion" id="recoje_direccion" maxlength="100" value="{{ selected_recoje_direccion or '' }}">
                    </div>
                    
                    <div class="form-group" id="agenciaGroup" style="display: none;">
                        <label for="agencia">Agencia:</label>
                        <select name="agencia" id="agencia" onchange="toggleOtraAgencia()">
                            <option value="">Seleccionar agencia</option>
                            <option value="FLORES HERMANOS" {% if selected_agencia=='FLORES HERMANOS' %}selected{% endif %}>FLORES HERMANOS</option>
                            <option value="MARVISUR" {% if selected_agencia=='MARVISUR' %}selected{% endif %}>MARVISUR</option>
                            <option value="ANTEZANA" {% if selected_agencia=='ANTEZANA' %}selected{% endif %}>ANTEZANA</option>
                            <option value="GRAEL" {% if selected_agencia=='GRAEL' %}selected{% endif %}>GRAEL</option>
                            <option value="RAZA" {% if selected_agencia=='RAZA' %}selected{% endif %}>RAZA</option>
                            <option value="RANA EXPRESS" {% if selected_agencia=='RANA EXPRESS' %}selected{% endif %}>RANA EXPRESS</option>
                            <option value="SHALOM" {% if selected_agencia=='SHALOM' %}selected{% endif %}>SHALOM</option>
                            <option value="TURISMOS DIAS" {% if selected_agencia=='TURISMOS DIAS' %}selected{% endif %}>TURISMOS DIAS</option>
                            <option value="OTRA" {% if selected_agencia=='OTRA' %}selected{% endif %}>Otra agencia</option>
                        </select>
                    </div>
                    
                    <div class="form-group" id="otraAgenciaGroup" style="display: none;">
                        <label for="otra_agencia">Especifique agencia:</label>
                        <textarea name="otra_agencia" id="otra_agencia" maxlength="60" rows="1">{{ selected_otra_agencia or '' }}</textarea>
                    </div>
                    
                    <div class="form-group" id="notasGroup" style="display: none;">
                        <label for="other_notes">Otras indicaciones:</label>
                        <textarea name="other_notes" id="other_notes" maxlength="120" rows="2">{{ selected_notes or '' }}</textarea>
                    </div>
                    
                    <button type="submit" class="btn btn-convert" id="convertirBtn">Convertir a PDF</button>
                    <br><br>
                    <a href="javascript:void(0)" class="btn btn-clean" onclick="confirmarLimpiar()">Limpiar</a>
                    {% if pdf_url %}
                    <a href="/download" class="btn btn-download">Descargar PDF</a>
                    {% endif %}
                </form>
            </div>
            
            <div class="left-panel">
                {% if pdf_url %}
                <iframe id="pdf-viewer" src="{{pdf_url}}" style="width:100%;height:600px;border:none;border-radius:8px;"></iframe>
                {% else %}
                <div class="empty-state">
                    <h2>Esperando documento</h2>
                    <p>Sube un archivo XML para previsualizar el PDF</p>
                </div>
                {% endif %}
                
                {% if info %}
                <div class="info">
                    <h3>Información del Documento</h3>
                    <p><strong>Tipo:</strong> {{ info.tipo }}</p>
                    <p><strong>Número:</strong> {{ info.numero }}</p>
                    <p><strong>Emisor:</strong> {{ info.emisor }}</p>
                    <p><strong>Cliente:</strong> {{ info.cliente }}</p>
                    <p><strong>Total:</strong> {{ info.total }}</p>
                    <p><strong>{% if info.tipo == 'YAPES RESUMEN' %}Rango de Fechas{% else %}Fecha{% endif %}:</strong> {{ info.fecha }}</p>
                </div>
                {% endif %}
            </div>
        </div>
    </div>
    
    <script>
    function updateFileName() {
        var input = document.getElementById('xml_file');
        var fileLabel = document.getElementById('fileLabel');
        
        if (input.files && input.files[0]) {
            var name = input.files[0].name;
            fileLabel.classList.add('active');
            fileLabel.innerHTML = name + ' <span onclick="removeFile(event)" style="cursor:pointer;font-size:1.2rem;">&times;</span>';
        }
    }
    
    function checkFormato() {
        toggleAgencia();
    }
    
    function validarArchivo() {
        var input = document.getElementById('xml_file');
        var fileLabel = document.getElementById('fileLabel');
        var formato = document.getElementById('formato').value;
        
        // Formato YAPES puede usar Google Sheets sin archivo
        if (formato === 'yapes' && (!input.files || !input.files[0]) && (!fileLabel.classList.contains('active'))) {
            return true;
        }
        
        if ((!input.files || !input.files[0]) && (!fileLabel.classList.contains('active'))) {
            Swal.fire({
                icon: 'warning',
                title: 'Archivo requerido',
                text: 'Por favor, ingresa un archivo XML o CSV válido para continuar',
                confirmButtonColor: '#667eea',
                confirmButtonText: 'Aceptar'
            });
            return false;
        }
        
        if (input.files && input.files[0]) {
            var fileName = input.files[0].name.toLowerCase();
            if (!fileName.endsWith('.xml') && !fileName.endsWith('.csv') && !fileName.endsWith('.xlsx')) {
                Swal.fire({
                    icon: 'error',
                    title: 'Archivo inválido',
                    text: 'El archivo debe ser XML, CSV o XLSX',
                    confirmButtonColor: '#dc2626',
                    confirmButtonText: 'Aceptar'
                });
                return false;
            }
        }
        
        // Muestra indicador de carga antes de enviar para evitar dobles clics
        Swal.fire({
            title: 'Procesando documento...',
            text: 'Por favor espera un momento',
            allowOutsideClick: false,
            showConfirmButton: false,
            didOpen: () => {
                Swal.showLoading();
            }
        });
        
        return true;
    }
    
    function confirmarLimpiar() {
        Swal.fire({
            icon: 'question',
            title: '¿Limpiar formulario?',
            text: 'Se eliminarán todos los datos ingresados',
            showCancelButton: true,
            confirmButtonColor: '#b91c1c',
            cancelButtonColor: '#64748b',
            confirmButtonText: 'Sí, limpiar',
            cancelButtonText: 'Cancelar'
        }).then((result) => {
            if (result.isConfirmed) {
                window.location.href = '/clear';
            }
        });
    }
    
    function toggleAgencia() {
        var formato = document.getElementById('formato').value;
        var agenciaGroup = document.getElementById('agenciaGroup');
        var notasGroup = document.getElementById('notasGroup');
        var recojeGroup = document.getElementById('recojeGroup');
        var yapesFechaGroup = document.getElementById('yapesFechaGroup');
        
        if (formato === 'shipping_label') {
            if(agenciaGroup) agenciaGroup.style.display = 'block';
            if(notasGroup) notasGroup.style.display = 'block';
            if(recojeGroup) recojeGroup.style.display = 'block';
            if(yapesFechaGroup) yapesFechaGroup.style.display = 'none';
        } else if (formato === 'yapes') {
            if(agenciaGroup) agenciaGroup.style.display = 'none';
            if(notasGroup) notasGroup.style.display = 'none';
            if(recojeGroup) recojeGroup.style.display = 'none';
            if(yapesFechaGroup) yapesFechaGroup.style.display = 'block';
        } else {
            if(agenciaGroup) agenciaGroup.style.display = 'none';
            if(notasGroup) notasGroup.style.display = 'none';
            if(recojeGroup) recojeGroup.style.display = 'none';
            if(yapesFechaGroup) yapesFechaGroup.style.display = 'none';
            
            if(document.getElementById('agencia')) document.getElementById('agencia').value = '';
            if(document.getElementById('otra_agencia')) document.getElementById('otra_agencia').value = '';
            if(document.getElementById('other_notes')) document.getElementById('other_notes').value = '';
            if(document.getElementById('recoje_otra_persona')) document.getElementById('recoje_otra_persona').checked = false;
            
            toggleRecojeOtraPersona();
            toggleOtraAgencia();
        }
    }
    
    function toggleRecojeOtraPersona() {
        var checkbox = document.getElementById('recoje_otra_persona');
        var datosGroup = document.getElementById('recojeDatosGroup');
        
        if (checkbox && datosGroup) {
            if (checkbox.checked) {
                datosGroup.style.display = 'block';
            } else {
                datosGroup.style.display = 'none';
                if(document.getElementById('recoje_dni')) document.getElementById('recoje_dni').value = '';
                if(document.getElementById('recoje_nombre')) document.getElementById('recoje_nombre').value = '';
                if(document.getElementById('recoje_direccion')) document.getElementById('recoje_direccion').value = '';
            }
        }
    }
    
    function toggleOtraAgencia() {
        var agencia = document.getElementById('agencia').value;
        var otraAgenciaGroup = document.getElementById('otraAgenciaGroup');
        
        if (otraAgenciaGroup) {
            if (agencia === 'OTRA') {
                otraAgenciaGroup.style.display = 'block';
            } else {
                otraAgenciaGroup.style.display = 'none';
                if(document.getElementById('otra_agencia')) document.getElementById('otra_agencia').value = '';
            }
        }
    }
    
    function removeFile(e) {
        if (e) e.stopPropagation();
        var input = document.getElementById('xml_file');
        var fileLabel = document.getElementById('fileLabel');
        
        input.value = '';
        fileLabel.classList.remove('active');
        fileLabel.innerHTML = 'Subir archivo';
    }

    function inicializarInterfaz() {
        // Asegurarnos de que SweetAlert cierre cualquier carga colgada en el caché
        if (typeof Swal !== 'undefined' && Swal.close) {
            Swal.close();
        }

        // Ejecutar los toggles para sincronizar la UI según la opción cargada
        toggleAgencia();
        toggleRecojeOtraPersona();
        toggleOtraAgencia();

        {% if xml_file_name %}
        var fileLabel = document.getElementById('fileLabel');
        if (fileLabel) {
            fileLabel.classList.add('active');
            fileLabel.innerHTML = '{{ xml_file_name }} <span onclick="removeFile(event)" style="cursor:pointer;font-size:1.2rem;">&times;</span>';
        }
        
        var formato = '{{ selected_formato }}';
        if (formato === 'shipping_label') {
            var agencia = '{{ selected_agencia }}';
            if (agencia === 'OTRA' && document.getElementById('otraAgenciaGroup')) {
                document.getElementById('otraAgenciaGroup').style.display = 'block';
            }
            
            var recoje = '{{ selected_recoje }}';
            if (recoje === 'true') {
                if(document.getElementById('recoje_otra_persona')) document.getElementById('recoje_otra_persona').checked = true;
                if(document.getElementById('recojeDatosGroup')) document.getElementById('recojeDatosGroup').style.display = 'block';
            }
        }
        {% endif %}
    }

    // Sincronizaciones seguras en ciclos de vida de la pestaña
    document.addEventListener('DOMContentLoaded', inicializarInterfaz);
    window.addEventListener('pageshow', inicializarInterfaz);
    </script>
</body>
</html>
"""


@app.route('/')
@log_request
def index():
    """Página principal"""
    temp_id = session.get('current_pdf')
    pdf_url = None
    info = None
    pdf_name = None
    xml_file_name = None
    selected_formato = 'ticket'
    selected_agencia = ''
    selected_otra_agencia = ''
    selected_notes = ''
    selected_recoje = ''
    selected_recoje_dni = ''
    selected_recoje_nombre = ''
    selected_recoje_direccion = ''
    selected_yapes_fecha_inicio = ''
    selected_yapes_fecha_fin = ''
    
    if temp_id:
        pdf_path = os.path.join('temp_files', f'{temp_id}.pdf')
        if os.path.exists(pdf_path):
            pdf_url = url_for('view_pdf', temp_id=temp_id)
            info = session.get('pdf_info')
            pdf_name = session.get('pdf_name')
            xml_file_name = session.get('xml_file_name')
            selected_formato = session.get('selected_formato', 'ticket')
            selected_agencia = session.get('selected_agencia', '')
            selected_otra_agencia = session.get('selected_otra_agencia', '')
            selected_notes = session.get('selected_notes', '')
            selected_recoje = session.get('selected_recoje', '')
            selected_recoje_dni = session.get('selected_recoje_dni', '')
            selected_recoje_nombre = session.get('selected_recoje_nombre', '')
            selected_recoje_direccion = session.get('selected_recoje_direccion', '')
            selected_yapes_fecha_inicio = session.get('selected_yapes_fecha_inicio', '')
            selected_yapes_fecha_fin = session.get('selected_yapes_fecha_fin', '')
    
    return render_template_string(HTML_TEMPLATE, pdf_url=pdf_url, info=info, pdf_name=pdf_name, xml_file_name=xml_file_name, selected_formato=selected_formato, selected_agencia=selected_agencia, selected_otra_agencia=selected_otra_agencia, selected_notes=selected_notes, selected_recoje=selected_recoje, selected_recoje_dni=selected_recoje_dni, selected_recoje_nombre=selected_recoje_nombre, selected_recoje_direccion=selected_recoje_direccion, selected_yapes_fecha_inicio=selected_yapes_fecha_inicio, selected_yapes_fecha_fin=selected_yapes_fecha_fin)


@app.route('/convertir', methods=['GET', 'POST'])
@log_request
def convertir():
    """Endpoint para convertir XML/CSV a PDF"""
    try:
        # Si es GET, redirigir al home
        if request.method == 'GET':
            return redirect(url_for('index'))
        
        # Obtener datos del formulario
        xml_file = request.files.get('xml_file')
        formato = request.form.get('formato', CONFIG['DEFAULT_FORMAT'])
        agencia = request.form.get('agencia', '')
        otra_agencia = request.form.get('otra_agencia', '')
        other_notes = request.form.get('other_notes', '')
        recoje_otra_persona = request.form.get('recoje_otra_persona', '')
        recoje_dni = request.form.get('recoje_dni', '')
        recoje_nombre = request.form.get('recoje_nombre', '')
        recoje_direccion = request.form.get('recoje_direccion', '')
        yapes_fecha_inicio = request.form.get('yapes_fecha_inicio', '')
        yapes_fecha_fin = request.form.get('yapes_fecha_fin', '')
        logger.info(f"YAPES INPUT fechas: inicio={yapes_fecha_inicio}, fin={yapes_fecha_fin}")
        
        # Verificar si hay archivo nuevo o usar el de sesión
        xml_data = None
        filename = ''
        xml_path = ''
        
        if xml_file and xml_file.filename:
            # Archivo nuevo seleccionado
            import uuid as uuid_module
            xml_data = xml_file.read()
            filename = xml_file.filename.lower()
            
            logger.info(f"Archivo recibido: {filename}, tamaño={len(xml_data)} bytes")
            
            # Guardar en sesión con Redis (persiste entre requests)
            session['xml_file_data'] = xml_data
            session['xml_file_name'] = xml_file.filename
            session['selected_formato'] = formato
            session.modified = True
            
            # Verificar que se guardó
            saved_data = session.get('xml_file_data')
            logger.info(f"Guardando XML en sesión: {len(xml_data)} bytes, verificado: {len(saved_data) if saved_data else 0} bytes")
            
            # También guardar en archivo para proceso actual
            os.makedirs('temp_files', exist_ok=True)
            xml_temp_id = uuid_module.uuid4().hex
            xml_path = os.path.join('temp_files', f'{xml_temp_id}.xml')
            with open(xml_path, 'wb') as f:
                f.write(xml_data)
            session['xml_file_path'] = xml_path
        else:
            # Usar archivo existente de sesión
            xml_path = session.get('xml_file_path', '')
            filename = session.get('xml_file_name', '').lower()
            # Primero intentar usar datos de sesión (con Redis esto funciona)
            xml_data = session.get('xml_file_data')
            logger.info(f"Recuperando de sesión: xml_file_data={len(xml_data) if xml_data else 'None'} bytes")
            if not xml_data and xml_path and os.path.exists(xml_path):
                with open(xml_path, 'rb') as f:
                    xml_data = f.read()
                logger.info(f"Recuperando de archivo: {xml_path}={len(xml_data)} bytes")
            logger.info(f"Usando datos: {len(xml_data) if xml_data else 0} bytes")
        
        if not xml_data and formato != 'yapes':
            return render_template_string(HTML_TEMPLATE, error="Por favor, selecciona un archivo")
        
        # Construir agency_name
        agency_name = otra_agencia if agencia == 'OTRA' else agencia
        
        # Determinar tipo de archivo
        if filename.endswith(('.csv', '.xlsx')) or formato == 'yapes':
            ext = '.csv'
            if filename.endswith('.xlsx'):
                ext = '.xlsx'
            elif not filename.endswith(('.csv', '.xlsx')) and xml_data:
                return render_template_string(HTML_TEMPLATE, error="Para formato YAPES debe subir un archivo CSV o XLSX")
            
            import uuid as uuid_module
            pdf_temp_id = uuid_module.uuid4().hex
            output_path = os.path.join('temp_files', f'{pdf_temp_id}.pdf')
            
            # Si no hay archivo subido, usar Google Sheets
            if not xml_data and YAPES_SHEETS_URL:
                yapes = YapesPDF(output_path=output_path, fecha_inicio=yapes_fecha_inicio, fecha_fin=yapes_fecha_fin, sheets_url=YAPES_SHEETS_URL)
                if not yapes.parse():
                    return render_template_string(HTML_TEMPLATE, error="Error al conectar con Google Sheets")
                if not yapes.data:
                    inicio_fmt = datetime.strptime(yapes_fecha_inicio, '%Y-%m-%d').strftime('%d/%m/%Y') if yapes_fecha_inicio else 'N/A'
                    fin_fmt = datetime.strptime(yapes_fecha_fin, '%Y-%m-%d').strftime('%d/%m/%Y') if yapes_fecha_fin else 'N/A'
                    msg = f"No se encontraron registros entre {inicio_fmt} y {fin_fmt} en Google Sheets."
                    return render_template_string(HTML_TEMPLATE, error=msg)
                yapes.generate_pdf()
                pdf_name = f"YAPES_RESUMEN_{datetime.now().strftime('%Y%m%d')}.pdf"
                info = {
                    'tipo': 'YAPES RESUMEN',
                    'numero': f"{len(yapes.data)} registros",
                    'emisor': '-',
                    'cliente': 'Varios',
                    'total': '-',
                    'fecha': yapes.rango_fechas_yape if yapes.rango_fechas_yape else datetime.now().strftime('%d/%m/%Y')
                }
                temp_id = pdf_temp_id
                session['current_pdf'] = temp_id
                session['pdf_name'] = pdf_name
                session['pdf_info'] = info
                session['xml_file_name'] = session.get('xml_file_name', '')
                session['selected_formato'] = formato
                session['selected_agencia'] = ''
                session['selected_otra_agencia'] = ''
                session['selected_notes'] = ''
                session['selected_recoje'] = ''
                session['selected_recoje_dni'] = ''
                session['selected_recoje_nombre'] = ''
                session['selected_recoje_direccion'] = ''
                session['yapes_sheets_url'] = YAPES_SHEETS_URL
                if yapes_fecha_inicio:
                    session['selected_yapes_fecha_inicio'] = yapes_fecha_inicio
                if yapes_fecha_fin:
                    session['selected_yapes_fecha_fin'] = yapes_fecha_fin
                session.modified = True
                pdf_url = url_for('view_pdf', temp_id=temp_id)
                logger.info(f"PDF generado desde Google Sheets: temp_id={temp_id}, url={pdf_url}")
                return redirect(url_for('index'))
            
            # Procesar archivo subido para YAPES
                import uuid as uuid_module
                csv_temp_id = uuid_module.uuid4().hex
                csv_path = os.path.join('temp_files', f'{csv_temp_id}{ext}')
                
                if not xml_data:
                    xml_data = session.get('xml_file_data')
                
                if isinstance(xml_data, str):
                    xml_data = xml_data.encode('utf-8')
                
                with open(csv_path, 'wb') as f:
                    f.write(xml_data)
                
                logger.info(f"Guardando archivo en disco: {len(xml_data)} bytes, verificado: {os.path.getsize(csv_path)} bytes")
                session['csv_file_path'] = csv_path
                
                yapes = YapesPDF(csv_path, output_path, yapes_fecha_inicio, yapes_fecha_fin)
            
            if not yapes.parse():
                return render_template_string(HTML_TEMPLATE, error="Error al procesar el archivo. Formato esperado: Nombre,Monto,Fecha")
            
            if not yapes.data:
                inicio_fmt = datetime.strptime(yapes_fecha_inicio, '%Y-%m-%d').strftime('%d/%m/%Y') if yapes_fecha_inicio else 'N/A'
                fin_fmt = datetime.strptime(yapes_fecha_fin, '%Y-%m-%d').strftime('%d/%m/%Y') if yapes_fecha_fin else 'N/A'
                msg = f"No se encontraron registros entre {inicio_fmt} y {fin_fmt}. Verifique que el archivo contenga datos en ese rango."
                logger.warning(f"YAPES sin datos: {msg}")
                return render_template_string(HTML_TEMPLATE, error=msg)
            
            yapes.generate_pdf()
            
            # Nombre del archivo PDF
            fecha = datetime.now().strftime('%Y%m%d')
            pdf_name = f"YAPES_RESUMEN_{fecha}.pdf"
            
            info = {
                'tipo': 'YAPES RESUMEN',
                'numero': f"{len(yapes.data)} registros",
                'emisor': '-',
                'cliente': 'Varios',
                'total': '-',
                'fecha': yapes.rango_fechas_yape if yapes.rango_fechas_yape else datetime.now().strftime('%d/%m/%Y')
            }
            
        elif filename.endswith('.xml'):
            # Procesar XML para Ticket
            if not filename.endswith('.xml'):
                return render_template_string(HTML_TEMPLATE, error="El archivo debe ser de tipo XML (.xml)")
            
            # Usar el xml_path local
            if not xml_path:
                xml_path = session.get('xml_file_path', '')
            
            logger.info(f"Procesando XML: xml_path={xml_path}, existe={os.path.exists(xml_path) if xml_path else False}")
            
            pdf_temp_id = uuid.uuid4().hex
            os.makedirs('temp_files', exist_ok=True)
            output_path = os.path.join('temp_files', f'{pdf_temp_id}.pdf')
            
            extra_data = {}
            if formato == 'shipping_label':
                extra_data = {
                    'agency_name': agency_name,
                    'other_notes': other_notes,
                    'recoje_otra_persona': recoje_otra_persona == 'on',
                    'recoje_dni': recoje_dni,
                    'recoje_nombre': recoje_nombre.upper() if recoje_nombre else '',
                    'recoje_direccion': recoje_direccion.upper() if recoje_direccion else ''
                }
            
            factura = FacturaXMLtoPDF(xml_path, output_path, extra_data)
            
            if not factura.parse_xml():
                error_msg = factura.errors[0] if factura.errors else "Error al procesar el XML"
                return render_template_string(HTML_TEMPLATE, error=f"{error_msg}")
                        
            factura.generate_pdf(formato)
            logger.info(f"PDF generado exitosamente: {output_path}, existe={os.path.exists(output_path)}")
            numero = factura.data.get('numero_factura', 'documento')
            cliente_nombre = factura.data.get('cliente_nombre', 'cliente').replace(' ', '_')
            cliente_nombre = ''.join(c for c in cliente_nombre if c.isalnum() or c == '_')
            pdf_name = f"{numero}_{cliente_nombre}_{formato}.pdf"
            
            info = {
                'tipo': factura.data.get('tipo_documento', 'N/A'),
                'numero': numero,
                'emisor': factura.data.get('emisor_nombre', 'N/A'),
                'cliente': factura.data.get('cliente_nombre', 'N/A'),
                'total': factura.format_currency(factura.data.get('total_pagar', '0.00')),
                'fecha': factura.data.get('fecha_emision', 'N/A')
            }
        else:
            return render_template_string(HTML_TEMPLATE, error="Formato no soportado. Use XML, CSV o XLSX")
        
        # Guardar PDF en sesión (en memoria para Render)
        temp_id = pdf_temp_id
        session['current_pdf'] = temp_id
        session['pdf_name'] = pdf_name
        session['pdf_info'] = info
        session['xml_file_name'] = session.get('xml_file_name', '')
        session['selected_formato'] = formato
        session['selected_agencia'] = agencia if formato == 'shipping_label' else ''
        session['selected_otra_agencia'] = otra_agencia if formato == 'shipping_label' else ''
        session['selected_notes'] = other_notes if formato == 'shipping_label' else ''
        session['selected_recoje'] = 'true' if recoje_otra_persona == 'on' else ''
        session['selected_recoje_dni'] = recoje_dni if formato == 'shipping_label' else ''
        session['selected_recoje_nombre'] = recoje_nombre if formato == 'shipping_label' else ''
        session['selected_recoje_direccion'] = recoje_direccion if formato == 'shipping_label' else ''
        if yapes_fecha_inicio:
            session['selected_yapes_fecha_inicio'] = yapes_fecha_inicio
        if yapes_fecha_fin:
            session['selected_yapes_fecha_fin'] = yapes_fecha_fin
        session.modified = True
        
        # Retornar con vista previa
        pdf_url = url_for('view_pdf', temp_id=temp_id)
        logger.info(f"PDF generado: temp_id={temp_id}, url={pdf_url}, xml_data={session.get('xml_file_data') is not None}")
        return redirect(url_for('index'))        
    except Exception as e:
        logger.exception("Error en conversión")
        return render_template_string(HTML_TEMPLATE, error=f"Error: {str(e)}")


@app.route('/health')
def health():
    """Health check"""
    return jsonify({'status': 'ok', 'service': 'XML to PDF Converter'})

@app.route('/images/<path:filename>')
def serve_image(filename):
    """Servir archivos de la carpeta images"""
    from flask import send_from_directory
    return send_from_directory('images', filename)


@app.route('/pdf/<temp_id>')
def view_pdf(temp_id):
    """Servir PDF - regenera desde sesión si es necesario"""
    try:
        # Verificar que el temp_id coincida
        if temp_id != session.get('current_pdf'):
            return "PDF no encontrado o expirado", 404
        
        # Obtener datos de sesión
        xml_data = session.get('xml_file_data')
        formato = session.get('selected_formato', 'ticket')
        
        if not xml_data and formato != 'yapes':
            return "Archivo no encontrado. Por favor, sube el archivo nuevamente.", 404
        
        # Generar PDF temporal
        os.makedirs('temp_files', exist_ok=True)
        import uuid as uuid_module
        pdf_temp_id = uuid_module.uuid4().hex
        output_path = os.path.join('temp_files', f'{pdf_temp_id}.pdf')
        
        if not xml_data and formato == 'yapes':
            xml_path = ''
        else:
            xml_temp_id = uuid_module.uuid4().hex
            xml_path = os.path.join('temp_files', f'{xml_temp_id}.xml')
            with open(xml_path, 'wb') as f:
                f.write(xml_data if isinstance(xml_data, bytes) else xml_data.encode('utf-8'))
        
        # Obtener datos adicionales
        agencia = session.get('selected_agencia', '')
        otra_agencia = session.get('selected_otra_agencia', '')
        other_notes = session.get('selected_notes', '')
        recoje_otra_persona = session.get('selected_recoje', '') == 'true'
        recoje_dni = session.get('selected_recoje_dni', '')
        recoje_nombre = session.get('selected_recoje_nombre', '')
        recoje_direccion = session.get('selected_recoje_direccion', '')
        
        agency_name = otra_agencia if agencia == 'OTRA' else agencia
        
        extra_data = {}
        if formato == 'shipping_label':
            extra_data = {
                'agency_name': agency_name,
                'other_notes': other_notes,
                'recoje_otra_persona': recoje_otra_persona,
                'recoje_dni': recoje_dni,
                'recoje_nombre': recoje_nombre.upper() if recoje_nombre else '',
                'recoje_direccion': recoje_direccion.upper() if recoje_direccion else ''
            }
        
        if formato == 'yapes':
            sheets_url = session.get('yapes_sheets_url', '')
            csv_path = session.get('csv_file_path')
            ext = session.get('yapes_file_ext', '.csv')
            
            if sheets_url and (not csv_path or not os.path.exists(csv_path)):
                yapes_fecha_inicio = session.get('selected_yapes_fecha_inicio', '')
                yapes_fecha_fin = session.get('selected_yapes_fecha_fin', '')
                yapes = YapesPDF(output_path=output_path, fecha_inicio=yapes_fecha_inicio, fecha_fin=yapes_fecha_fin, sheets_url=sheets_url)
            else:
                if not csv_path or not os.path.exists(csv_path):
                    import uuid as uuid_module
                    csv_temp_id = uuid_module.uuid4().hex
                    csv_path = os.path.join('temp_files', f'{csv_temp_id}{ext}')
                    with open(csv_path, 'wb') as f:
                        f.write(xml_data if isinstance(xml_data, bytes) else xml_data.encode('utf-8'))
                    session['csv_file_path'] = csv_path
                yapes_fecha_inicio = session.get('selected_yapes_fecha_inicio', '')
                yapes_fecha_fin = session.get('selected_yapes_fecha_fin', '')
                yapes = YapesPDF(csv_path, output_path, yapes_fecha_inicio, yapes_fecha_fin)
            
            if not yapes.parse():
                return "Error al procesar el archivo", 500
            yapes.generate_pdf()
        else:
            factura = FacturaXMLtoPDF(xml_path, output_path, extra_data)
            if not factura.parse_xml():
                return "Error al procesar el XML", 500
            factura.generate_pdf(formato)
        
        logger.info(f"PDF regenerado desde sesión: {output_path}")
        return send_file(output_path, mimetype='application/pdf', as_attachment=False)
    except Exception as e:
        logger.exception("Error sirviendo PDF")
        return f"Error: {str(e)}", 500
        
    except Exception as e:
        logger.exception("Error sirviendo PDF")
        return f"Error: {str(e)}", 500


@app.route('/download')
def download_pdf():
    """Descargar PDF - regenera desde sesión si es necesario"""
    try:
        temp_id = session.get('current_pdf')
        if not temp_id:
            return "PDF no encontrado", 404
        
        # Obtener datos de sesión
        xml_data = session.get('xml_file_data')
        pdf_name = session.get('pdf_name', 'documento.pdf')
        formato = session.get('selected_formato', 'ticket')
        
        if not xml_data and formato != 'yapes':
            return "Archivo no encontrado. Por favor, sube el archivo nuevamente.", 404
        
        # Generar PDF temporal
        os.makedirs('temp_files', exist_ok=True)
        import uuid as uuid_module
        pdf_temp_id = uuid_module.uuid4().hex
        output_path = os.path.join('temp_files', f'{pdf_temp_id}.pdf')
        
        if not xml_data and formato == 'yapes':
            xml_path = ''
        else:
            xml_temp_id = uuid_module.uuid4().hex
            xml_path = os.path.join('temp_files', f'{xml_temp_id}.xml')
            with open(xml_path, 'wb') as f:
                f.write(xml_data if isinstance(xml_data, bytes) else xml_data.encode('utf-8'))
        
        # Obtener datos adicionales
        agencia = session.get('selected_agencia', '')
        otra_agencia = session.get('selected_otra_agencia', '')
        other_notes = session.get('selected_notes', '')
        recoje_otra_persona = session.get('selected_recoje', '') == 'true'
        recoje_dni = session.get('selected_recoje_dni', '')
        recoje_nombre = session.get('selected_recoje_nombre', '')
        recoje_direccion = session.get('selected_recoje_direccion', '')
        
        agency_name = otra_agencia if agencia == 'OTRA' else agencia
        
        extra_data = {}
        if formato == 'shipping_label':
            extra_data = {
                'agency_name': agency_name,
                'other_notes': other_notes,
                'recoje_otra_persona': recoje_otra_persona,
                'recoje_dni': recoje_dni,
                'recoje_nombre': recoje_nombre.upper() if recoje_nombre else '',
                'recoje_direccion': recoje_direccion.upper() if recoje_direccion else ''
            }
        
        if formato == 'yapes':
            sheets_url = session.get('yapes_sheets_url', '')
            csv_path = session.get('csv_file_path')
            ext = session.get('yapes_file_ext', '.csv')
            
            if sheets_url and (not csv_path or not os.path.exists(csv_path)):
                yapes_fecha_inicio = session.get('selected_yapes_fecha_inicio', '')
                yapes_fecha_fin = session.get('selected_yapes_fecha_fin', '')
                yapes = YapesPDF(output_path=output_path, fecha_inicio=yapes_fecha_inicio, fecha_fin=yapes_fecha_fin, sheets_url=sheets_url)
            else:
                if not csv_path or not os.path.exists(csv_path):
                    import uuid as uuid_module
                    csv_temp_id = uuid_module.uuid4().hex
                    csv_path = os.path.join('temp_files', f'{csv_temp_id}{ext}')
                    with open(csv_path, 'wb') as f:
                        f.write(xml_data if isinstance(xml_data, bytes) else xml_data.encode('utf-8'))
                    session['csv_file_path'] = csv_path
                yapes_fecha_inicio = session.get('selected_yapes_fecha_inicio', '')
                yapes_fecha_fin = session.get('selected_yapes_fecha_fin', '')
                yapes = YapesPDF(csv_path, output_path, yapes_fecha_inicio, yapes_fecha_fin)
            
            if not yapes.parse():
                return "Error al procesar el archivo", 500
            yapes.generate_pdf()
        else:
            factura = FacturaXMLtoPDF(xml_path, output_path, extra_data)
            if not factura.parse_xml():
                return "Error al procesar el XML", 500
            factura.generate_pdf(formato)
        
        return send_file(output_path, mimetype='application/pdf', as_attachment=True, download_name=pdf_name)
    except Exception as e:
        logger.exception("Error descargando PDF")
        return str(e), 500


@app.route('/clear')
def clear_session():
    """Limpiar sesión y archivos temporales"""
    try:
        # Limpiar archivos temporales
        temp_id = session.get('current_pdf')
        if temp_id:
            pdf_path = session.get(f'pdf_{temp_id}')
            if pdf_path and os.path.exists(pdf_path):
                try:
                    os.unlink(pdf_path)
                except:
                    pass
            # Limpiar claves de sesión
            session.pop(f'pdf_{temp_id}', None)
        
        session.pop('current_pdf', None)
        session.pop('pdf_name', None)
        session.pop('pdf_info', None)
        
        return redirect(url_for('index'))
    except Exception as e:
        logger.exception("Error limpiando sesión")
        return redirect(url_for('index'))


# ========== MAIN ==========
if __name__ == '__main__':
    os.makedirs("images", exist_ok=True)
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)